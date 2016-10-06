from bs4 import BeautifulSoup
from urlparse import urlparse
import requests
import time
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('Rank.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

pages_deep = 1

def google_rank_checker(query,i,t): #query is the keyword for which you want to track the rank of websites
	number = 0    # keeps track of the ranking position
    
	for start in range(int(pages_deep)):
		url = "https://www.google.co.in/search?q="+query.replace(' ','+')+"&oq="+query.replace(' ','+')+"&aqs=chrome.0.69i59j69i61l3j69i59.6617j0j4&sourceid=chrome&ie=UTF-8"
        r = requests.get(url)
        data = r.text
        soup = BeautifulSoup(data,"lxml")
        
        for result in soup.find_all(id='search')[0].find_all('cite'):
            number += 1
            domain = result.text
            domain = domain.replace('https://', '')
            if "Website1" in domain: #enter name of website1
            	sheet.cell(row=i+1,column=t).value = number
            if "Website2" in domain: #enter name of website2
            	sheet.cell(row=i+2,column=t).value = number
            if "Website3" in domain: #enter name of website3
            	sheet.cell(row=i+3,column=t).value = number
            if "Website4" in domain: #enter name of website4 and increase further if you want to track more websites
            	sheet.cell(row=i+4,column=t).value = number
            print number, domain
t=2
while t<5:
	sheet.cell(row=1,column=t).value = datetime.now().time()
	for i in range(2,25,6):
		query=sheet.cell(row=i,column=1).value
		google_rank_checker(query,i,t)
	t=t+1
	time.sleep(60)

wb.save('Rank.xlsx')